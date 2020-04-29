import openpyxl
import pymysql
import re


dest_path = "test.xlsx"

wb = openpyxl.Workbook()

connection = pymysql.connect(host='172.31.47.159', 
                user='icity_dev', 
                password='icity_dev', 
                db='omp_dev', 
                charset='utf8mb4', 
                cursorclass=pymysql.cursors.DictCursor)


head = ["序号", "属性名称",	"属性编码",	"属性描述",	"数据类型",	"字段长度",	"字段格式",	"主键标识",	"非空标识",	"对应主数据",	"系统来源",	"校验规则"]

def write_head(ws):
    for col in range(len(head)):
        ws.cell(row=1, column=col+1, value=head[col])

def write_data(ws, connection, table_name):
    write_head(ws)
    with connection.cursor() as cursor:
        cursor.execute(f"show full fields from {table_name};")
        fields = cursor.fetchall()
        row_num = len(fields)
        for row in range(row_num):
            ws.cell(row=row+2, column=1, value=row+1)
            ws.cell(row=row+2, column=2, value=fields[row]['Comment'])
            ws.cell(row=row+2, column=3, value=fields[row]['Field'])
            type = fields[row]['Type']
            ws.cell(row=row+2, column=5, value=re.search("[a-z]+", type).group())
            ret = re.search("\d+", type)
            if ret != None:
                ws.cell(row=row+2, column=6, value=ret.group())
            if fields[row]['Key'] == 'PRI':
                ws.cell(row=row+2, column=8, value="是")
            else:
                ws.cell(row=row+2, column=8, value="否")
            if fields[row]['Null'] == 'NO':
                ws.cell(row=row+2, column=9, value="非空")
            else:
                ws.cell(row=row+2, column=9, value="可空")



try:
    with connection.cursor() as cursor:
        sql = "show tables;"
        cursor.execute(sql)
        result = cursor.fetchall()
        index = 0
        for table in result:
            table_name = table['Tables_in_omp_dev']
            if index == 0:
                ws = wb.active
                ws.title = table_name
            else:
                ws = wb.create_sheet(title=table_name)
            write_head(ws)
            write_data(ws, connection, table_name)
            index+=1
finally:
    connection.close()

wb.save(dest_path)



